import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import matplotlib
import requests

# --- CRITICAL: mplcairo Initialization for Devanagari Font Shaping ---
try:
    import mplcairo  # noqa: F401
    matplotlib.use("module://mplcairo.tk")
except ImportError:
    matplotlib.use("TkAgg")

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.font_manager as fm
import matplotlib.colors as mcolors

# Cross-platform winsound shim (no-op on non-Windows)
try:
    import winsound
    def _beep():
        try:
            winsound.MessageBeep(winsound.MB_ICONASTERISK)
        except Exception:
            pass
except Exception:
    def _beep():
        pass

import sys
import os
import ctypes
import re
from io import BytesIO
from scraper import fetch_pr_votes
import numpy as np
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
from pr_excel_helper import update_excel_from_web, read_pr_from_excel

# --- Global Configuration ---
FONT_NAME = "Arial"
MAROON_COLOR = "#800000"
PRIMARY_TEXT = "#2c3e50"
SUBTLE_TEXT = "#5d6d7e"
ACCENT_BLUE = "#2980b9"

matplotlib.rcParams["axes.unicode_minus"] = False

REQUEST_TIMEOUT = 10


def make_white_transparent(img):
    img = img.convert("RGBA")
    datas = img.getdata()
    newData = []
    for item in datas:
        if item[0] > 220 and item[1] > 220 and item[2] > 220:
            newData.append((255, 255, 255, 0))
        else:
            newData.append(item)
    img.putdata(newData)
    return img


def get_best_font_color(hex_color):
    rgb = mcolors.to_rgb(hex_color)
    luminance = (0.2126 * rgb[0] + 0.7152 * rgb[1] + 0.0722 * rgb[2])
    return 'white' if luminance < 0.55 else MAROON_COLOR


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  # type: ignore[attr-defined]
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def calculate_seats(df, total_seats, threshold_pct, manual_total_votes):
    """
    Returns:
        (result_df_sorted,
         total_input_votes,
         threshold_limit,
         qualified_count,
         total_input_parties,
         total_valid_votes)   # sum of votes of qualified parties
    """
    threshold_limit = manual_total_votes * (threshold_pct / 100)
    eligible_df = df[df["Votes"] >= threshold_limit].copy()
    total_input_parties = len(df)
    if eligible_df.empty:
        return None, manual_total_votes, threshold_limit, 0, total_input_parties, 0.0

    total_valid_votes = float(eligible_df["Votes"].sum())

    eligible_df["Seats"] = 0
    eligible_df["Votes"] = eligible_df["Votes"].astype(float)

    for _ in range(total_seats):
        eligible_df["Divisor"] = eligible_df["Seats"].apply(
            lambda s: 1.4 if s == 0 else (2 * s + 1)
        )
        eligible_df["Quotient"] = eligible_df["Votes"] / eligible_df["Divisor"]
        winner_idx = eligible_df["Quotient"].idxmax()
        eligible_df.at[winner_idx, "Seats"] += 1

    return (
        eligible_df[["Party", "Votes", "Seats"]].sort_values(by="Seats", ascending=False),
        manual_total_votes,
        threshold_limit,
        len(eligible_df),
        total_input_parties,
        total_valid_votes,
    )


class ElectionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Nepal PR Seat Calculator Pro")
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # --- Initial window geometry ---
        w, h = 1280, 900
        ws, hs = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{int(ws/2 - w/2)}+{int(hs/2 - h/2)}")
        self.root.configure(bg="#f8f9fa")

        # --- App state ---
        self.result_df = None
        self.filepath = None
        self.current_figure = None
        self.total_input_votes = 0          # all votes
        self.total_valid_votes = 0          # sum of qualified parties' votes
        self.pan_start = None
        self.logo_cache = {}
        self.chart_logo_cache = {}
        self.table_images = []
        self.logo_map = {}
        self.data_source = ""

        # Network session
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "PumoriSeatCalc/1.0"})

        # --- Assets (logo/icon) ---
        logo_path = resource_path("pumori_logo.png")
        icon_path = resource_path("pumori_icon.ico")
        try:
            icon_img = ImageTk.PhotoImage(Image.open(icon_path))
            self.root.iconphoto(True, icon_img)
            self._icon_ref = icon_img

            load = Image.open(logo_path)
            self.logo_img = ImageTk.PhotoImage(load.resize((180, 50), Image.Resampling.LANCZOS))
        except Exception:
            self.logo_img = None

        # --- ttk styles ---
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Treeview", font=(FONT_NAME, 14), rowheight=35)
        style.configure("Treeview.Heading", font=(FONT_NAME, 14, "bold"), background="#d1d8e0")

        # ----- TOP-LEVEL GRID LAYOUT -----
        root.grid_rowconfigure(0, weight=0)
        root.grid_rowconfigure(1, weight=0)
        root.grid_rowconfigure(2, weight=1)
        root.grid_rowconfigure(3, weight=0)
        root.grid_columnconfigure(0, weight=1)

        # Header
        header = tk.Frame(root, bg="#2c3e50", height=80)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)

        if self.logo_img:
            tk.Label(header, image=self.logo_img, bg="#2c3e50").place(relx=0.0, rely=0.5, anchor="w", x=20)
        tk.Label(header, text="Nepal Election PR Seat Calculator", font=(FONT_NAME, 26, "bold"), fg="white", bg="#2c3e50").place(relx=0.5, rely=0.5, anchor="center")
        tk.Label(header, text="Developed By: Pumori Engineering Services", font=(FONT_NAME, 12), fg="#bdc3c7", bg="#2c3e50").place(relx=1.0, rely=0.5, anchor="e", x=-30)

        # CONTROL BAR (updated order)
        ctrl_bar = tk.Frame(root, bg="white", pady=12, highlightbackground="#e1e4e8", highlightthickness=1)
        ctrl_bar.grid(row=1, column=0, sticky="ew", padx=25, pady=10)

        btn_cfg = {"font": (FONT_NAME, 12, "bold"), "fg": "white", "padx": 15, "relief": "flat"}

        # 1) Run From Web
        tk.Button(ctrl_bar, text="🌐 Run From Web", command=self.run_from_web,
                bg="#16a085", **btn_cfg).grid(row=0, column=0, padx=10)

        # 2) Load Excel
        tk.Button(ctrl_bar, text="📁 Load Excel", command=self.load_file,
                bg="#34495e", **btn_cfg).grid(row=0, column=1, padx=10)

        # 3) Update Excel From Web
        tk.Button(ctrl_bar, text="🔄 Update Excel from Web", command=self.update_excel_from_web_btn,
                bg="#27ae60", **btn_cfg).grid(row=0, column=2, padx=10)

        # 4) Run From Excel
        tk.Button(ctrl_bar, text="▶ Run From Excel", command=self.run_from_excel_btn,
                bg="#006994", **btn_cfg).grid(row=0, column=3, padx=10)

        # 5) Total PR Seats textbox
        tk.Label(ctrl_bar, text="Total PR Seats:", bg="white", fg="#2c3e50",
                font=(FONT_NAME, 12, "bold")).grid(row=0, column=4, padx=(25, 5))
        self.seats_entry = tk.Entry(ctrl_bar, width=8, font=(FONT_NAME, 12), justify="center")
        self.seats_entry.insert(0, "110")
        self.seats_entry.grid(row=0, column=5, padx=5)

        # Body
        self.body_frame = tk.Frame(root, bg="#f8f9fa")
        self.body_frame.grid(row=2, column=0, sticky="nsew")

        self.main_container = tk.PanedWindow(self.body_frame, orient="horizontal", bg="#f8f9fa", sashwidth=6)
        self.main_container.pack(fill="both", expand=True, padx=25)
        self.main_container.bind("<B1-Motion>", lambda e: "break")

        # Left tables pane
        self.table_frame = tk.PanedWindow(self.main_container, orient="vertical", bg="white", sashwidth=4)
        self.main_container.add(self.table_frame, width=500, stretch="never")

        self.qualified_frame = tk.Frame(self.table_frame, bg="white")
        self.table_frame.add(self.qualified_frame, height=320)

        tk.Label(self.qualified_frame, text="Qualified Parties", font=(FONT_NAME, 14, "bold"), bg="white", fg=PRIMARY_TEXT).pack(anchor="w", padx=10, pady=(5, 2))

        self.tree_qualified = ttk.Treeview(self.qualified_frame, columns=("Party", "Votes", "Seats"), show="tree headings", height=8)
        self.tree_qualified.heading("#0", text="Symbol")
        self.tree_qualified.column("#0", width=60, anchor="center")
        self.tree_qualified.heading("Party", text="Political Party Name")
        self.tree_qualified.column("Party", width=200, anchor="w")
        self.tree_qualified.heading("Votes", text="Votes")
        self.tree_qualified.column("Votes", width=100, anchor="center")
        self.tree_qualified.heading("Seats", text="Seats")
        self.tree_qualified.column("Seats", width=60, anchor="center")
        self.tree_qualified.pack(fill="both", expand=True)

        # Unqualified
        self.unqualified_frame = tk.Frame(self.table_frame, bg="#f4f4f4")
        self.table_frame.add(self.unqualified_frame)
        tk.Label(self.unqualified_frame, text="Unqualified Parties", font=(FONT_NAME, 14, "bold"), bg="#f4f4f4", fg="#7f8c8d").pack(anchor="w", padx=10, pady=(5, 2))

        self.unqualified_canvas = tk.Canvas(self.unqualified_frame, bg="#f4f4f4", height=180, highlightthickness=0)
        self.unqualified_canvas.pack(side="left", fill="both", expand=True)
        scrollbar = tk.Scrollbar(self.unqualified_frame, orient="vertical", command=self.unqualified_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        self.unqualified_canvas.configure(yscrollcommand=scrollbar.set)

        self.unqualified_grid = tk.Frame(self.unqualified_canvas, bg="#f4f4f4")
        self.unqualified_canvas.create_window((0, 0), window=self.unqualified_grid, anchor="nw")
        self.unqualified_grid.bind("<Configure>", lambda e: self.unqualified_canvas.configure(scrollregion=self.unqualified_canvas.bbox("all")))

        # Right chart pane
        self.right_pane = tk.PanedWindow(self.main_container, orient="vertical", bg="#f8f9fa", sashwidth=6)
        self.main_container.add(self.right_pane)
        self.chart_frame = tk.Frame(self.right_pane, bg="white", bd=1, relief="solid")
        self.right_pane.add(self.chart_frame)

        # Summary (bottom, fixed) with ONLY: Total Votes | 3% Threshold | Total Valid Votes
        self.summary_frame = tk.Frame(root, bg="#ecf0f1", height=80)
        self.summary_frame.grid(row=3, column=0, sticky="ew", padx=25)
        self.summary_frame.grid_propagate(False)

        lbl_style = {"bg": "#ecf0f1", "font": (FONT_NAME, 14, "bold"), "fg": PRIMARY_TEXT}
        # ORDER: Total Votes, 3% Threshold, Total Valid Votes
        self.lbl_total_input_votes = tk.Label(self.summary_frame, text="Total Votes: -", **lbl_style)
        self.lbl_total_input_votes.pack(side="left", padx=40)

        self.lbl_threshold = tk.Label(self.summary_frame, text="3% Threshold: -", **lbl_style)
        self.lbl_threshold.pack(side="left", padx=40)

        self.lbl_total_votes = tk.Label(self.summary_frame, text="Total Eligible Votes: -", **lbl_style)
        self.lbl_total_votes.pack(side="left", padx=40)

        # Minimum size
        self.root.update_idletasks()
        try:
            fixed_h = header.winfo_height() + ctrl_bar.winfo_height() + self.summary_frame.winfo_height()
        except Exception:
            fixed_h = 80 + 60 + 80
        self.root.minsize(width=980, height=fixed_h + 120)

    def on_closing(self):
        try:
            if self.current_figure:
                plt.close(self.current_figure)
            plt.close("all")
            self.session.close()
        except Exception:
            pass
        self.root.destroy()
        os._exit(0)

    def clear_unqualified_grid(self):
        for widget in self.unqualified_grid.winfo_children():
            widget.destroy()

    def clean_number(self, text):
        if pd.isna(text):
            return 0.0
        text = str(text).strip().replace(",", "").replace("\xa0", "")
        text = re.sub(r"[^\d०१२३४५६७८९.]", "", text)
        translation_table = str.maketrans("०१२३४५६७८९", "0123456789")
        try:
            return float(text.translate(translation_table))
        except Exception:
            return 0.0

    def load_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if f:
            self.filepath = f
            _beep()

    def update_excel_from_web_btn(self):
        if not self.filepath:
            messagebox.showwarning("Select Excel", "Please select an Excel file first.")
            return
        try:
            result = update_excel_from_web(self.filepath)
            messagebox.showinfo("Update Complete",
                                f"Updated: {result['updated']}\nMissing: {len(result['not_found'])}")
        except Exception as e:
            messagebox.showerror("Update Failed", str(e))

    def run_from_excel_btn(self):
        if not self.filepath:
            messagebox.showwarning("Select Excel", "Please select an Excel file first.")
            return
        try:
            # 1) Read standardized data from Excel (Party, Votes, Logo) + total_votes
            df, total_votes = read_pr_from_excel(self.filepath)

            self.data_source = self.filepath

            # 2) Compute seats exactly like the web flow
            seats = int(self.seats_entry.get())
            base = df[["Party", "Votes"]].copy()
            res, _, thresh, _, _, total_valid = calculate_seats(base, seats, 3.0, float(total_votes))
            if res is None or res.empty:
                messagebox.showwarning("No Qualified Parties", "No parties met the threshold.")
                return

            # 3) Attach logos (already standardized column in df)
            logo_map = dict(zip(df["Party"], df["Logo"]))
            res["Logo"] = res["Party"].map(logo_map)
            res = res.sort_values(by="Votes", ascending=False)

            # 4) Update UI state
            self.result_df = res

            # Populate unqualified parties (same logic as web mode)
            all_parties = set(df["Party"])
            qualified = set(res["Party"])

            self.unqualified_df = (
                df[~df["Party"].isin(qualified)]
                .sort_values(by="Votes", ascending=False)
                .copy()
            )

            self.clear_unqualified_grid()

            row = 0
            col = 0

            for _, r in self.unqualified_df.iterrows():

                frame = tk.Frame(self.unqualified_grid, bg="#f4f4f4")
                frame.grid(row=row, column=col, padx=8, pady=3, sticky="w")

                logo_url = r.get("Logo")
                logo_img = None

                if isinstance(logo_url, str) and logo_url.strip():
                    try:
                        if logo_url not in self.logo_cache:
                            response = self.session.get(logo_url, timeout=REQUEST_TIMEOUT)
                            img = Image.open(BytesIO(response.content)).resize((20,20))
                            img = make_white_transparent(img)
                            self.logo_cache[logo_url] = ImageTk.PhotoImage(img)

                        logo_img = self.logo_cache[logo_url]

                    except Exception:
                        pass

                tk.Label(frame, image=logo_img, bg="#f4f4f4").pack(side="left")
                tk.Label(frame, text=f" ({int(r['Votes']):,})",
                        font=(FONT_NAME,11),
                        fg="#7f8c8d",
                        bg="#f4f4f4").pack(side="left")

                col += 1
                if col == 4:
                    col = 0
                    row += 1

            self.total_input_votes = float(total_votes)
            self.total_valid_votes = total_valid
            self.lbl_total_input_votes.config(text=f"Total Votes: {int(self.total_input_votes):,}")
            self.lbl_threshold.config(text=f"3% Threshold: {int(thresh):,}")
            self.lbl_total_votes.config(text=f"Total Valid Votes: {int(self.total_valid_votes):,}")

            # 5) Redraw table + chart
            self.refresh_table_and_chart()

            from openpyxl import load_workbook

            try:

                wb = load_workbook(self.filepath)

                ws_votes = wb["PR_votes"]
                ws_summary = wb["Summary"]

                # Find the seat column in summary
                seats_col = None
                for cell in ws_summary[1]:
                    if cell.value and "seat" in str(cell.value).lower():
                        seats_col = cell.col_idx

                if seats_col is None:
                    raise ValueError("Seats column not found in Summary sheet")

                # Map party → seats
                seat_map = dict(zip(self.result_df["Party"], self.result_df["Seats"]))

                # Loop through PR_votes rows
                for r in range(2, ws_votes.max_row + 1):

                    party = ws_votes.cell(row=r, column=2).value  # column B = Party

                    if party:
                        party = str(party).strip()

                        seats = seat_map.get(party, 0)

                        # Write to same row in Summary sheet
                        ws_summary.cell(row=r, column=seats_col).value = int(seats)

                wb.save(self.filepath)

            except Exception as e:
                print("Summary update error:", e)
        except Exception as e:
            messagebox.showerror("Excel Run Error", str(e))

    def run_from_web(self):
        try:
            self.table_images.clear()
            df, total_votes = fetch_pr_votes()
            self.data_source = "result.election.gov.np"

            self.logo_map = dict(zip(df["Party"], df.get("Logo", pd.Series([None]*len(df)))))
            self.total_input_votes = total_votes

            res, total_input, thresh, count, total_p, total_valid = calculate_seats(
                df[["Party", "Votes"]],
                int(self.seats_entry.get()),
                3.0,
                total_votes
            )

            if res is not None and "Logo" in df.columns:
                res = res.merge(df[["Party", "Logo"]], on="Party", how="left")
            self.result_df = res
            self.total_valid_votes = total_valid

            # Update summary labels
            self.lbl_total_input_votes.config(text=f"Total Votes: {int(self.total_input_votes):,}")
            self.lbl_threshold.config(text=f"3% Threshold: {int(thresh):,}")
            self.lbl_total_votes.config(text=f"Total Valid Votes: {int(self.total_valid_votes):,}")

            # Refresh qualified table
            for i in self.tree_qualified.get_children():
                self.tree_qualified.delete(i)
            if res is not None:
                for _, r in res.iterrows():
                    logo_url = r.get("Logo") if "Logo" in r else None
                    logo_img = None
                    if logo_url:
                        try:
                            if logo_url not in self.logo_cache:
                                response = self.session.get(logo_url, timeout=REQUEST_TIMEOUT)
                                img = Image.open(BytesIO(response.content)).convert("RGBA")
                                img = make_white_transparent(img)
                                self.logo_cache[logo_url] = ImageTk.PhotoImage(img.resize((24, 24)))
                                self.chart_logo_cache[logo_url] = np.array(img)
                            logo_img = self.logo_cache[logo_url]
                        except Exception:
                            pass
                    row_values = (
                        str(r["Party"]),
                        f"{int(r['Votes']):,}",
                        str(int(r["Seats"]))
                    )

                    self.tree_qualified.insert("", tk.END, text="", image=logo_img, values=row_values)
                    self.table_images.append(logo_img)

            # Unqualified grid
            self.clear_unqualified_grid()
            row = 0
            col = 0
            unqualified_df = df[~df["Party"].isin(set(res["Party"]) if res is not None else set())].copy()
            self.unqualified_df = unqualified_df
            for _, r in self.unqualified_df.iterrows():
                frame = tk.Frame(self.unqualified_grid, bg="#f4f4f4")
                frame.grid(row=row, column=col, padx=8, pady=3, sticky="w")
                logo_url = r.get("Logo") if "Logo" in r else None
                logo_img = None
                if logo_url:
                    try:
                        if logo_url not in self.logo_cache:
                            response = self.session.get(logo_url, timeout=REQUEST_TIMEOUT)
                            img = Image.open(BytesIO(response.content)).resize((20, 20))
                            img = make_white_transparent(img)
                            self.logo_cache[logo_url] = ImageTk.PhotoImage(img)
                        logo_img = self.logo_cache[logo_url]
                    except Exception:
                        pass
                tk.Label(frame, image=logo_img, bg="#f4f4f4").pack(side="left")
                tk.Label(frame, text=f" ({int(r['Votes']):,})", font=(FONT_NAME,11), fg="#7f8c8d", bg="#f4f4f4").pack(side="left")
                col += 1
                if col == 4:
                    col = 0
                    row += 1

            self.update_chart()
            _beep()
        except Exception as e:
            messagebox.showerror("Web Fetch Error", str(e))

    def refresh_table_and_chart(self):
        """Refresh table and chart when running from Excel."""

        # Clear qualified table
        for item in self.tree_qualified.get_children():
            self.tree_qualified.delete(item)

        self.table_images.clear()

        if self.result_df is not None and not self.result_df.empty:

            for _, r in self.result_df.iterrows():

                logo_img = None
                logo_url = r["Logo"] if "Logo" in r and pd.notna(r["Logo"]) else None

                if isinstance(logo_url, str) and logo_url.strip():
                    try:
                        if logo_url not in self.logo_cache:
                            resp = self.session.get(logo_url, timeout=10)
                            img = Image.open(BytesIO(resp.content)).convert("RGBA")
                            img = make_white_transparent(img)

                            # store full image for chart
                            self.chart_logo_cache[logo_url] = np.array(img)

                            # small version for table
                            table_img = img.resize((24, 24), Image.Resampling.LANCZOS)
                            self.logo_cache[logo_url] = ImageTk.PhotoImage(table_img)

                        logo_img = self.logo_cache[logo_url]

                    except Exception:
                        pass

                # Safe string conversion
                party = str(r["Party"])
                votes = f"{int(r['Votes']):,}"
                seats = str(int(r["Seats"]))

                row_values = (party, votes, seats)

                if logo_img is not None:
                    self.tree_qualified.insert(
                        "",
                        "end",
                        text="",
                        image=logo_img,
                        values=row_values
                    )
                else:
                    self.tree_qualified.insert(
                        "",
                        "end",
                        text="",
                        values=row_values
                    )

                self.table_images.append(logo_img)

        # Update chart
        self.update_chart()

    def update_chart(self):
        for w in self.chart_frame.winfo_children():
            w.destroy()
        if self.current_figure:
            try:
                plt.close(self.current_figure)
            except Exception:
                pass
        self.current_figure = None

        if self.result_df is None:
            return

        chart_data = self.result_df[self.result_df["Seats"] > 0].copy()
        if chart_data.empty:
            return

        fig, ax = plt.subplots(figsize=(10, 8), dpi=120, facecolor="white")
        self.current_figure = fig

        total_seats = int(chart_data["Seats"].sum())
        colors = list(matplotlib.colormaps.get_cmap("tab20").colors)
        slice_colors = [
            "#87CEEB",  # Sky blue
            "#228B22",  # Tree green
            "#8B0000",  # Dark red
            "#FF6F6F",  # Light red
            "#8B5A2B",  # Mud
            "#A0522D"   # Wood
        ]

        slice_colors = (slice_colors * 10)[:len(chart_data)]

        def autopct_func(pct):
            seats = int(round(pct * total_seats / 100.0))
            return f"{pct:.1f}%\n{seats}"

        prop = fm.FontProperties(family=FONT_NAME)
        wedges, texts, autotexts = ax.pie(
            chart_data["Seats"], labels=None, autopct=autopct_func, startangle=90,
            counterclock=False, colors=slice_colors, labeldistance=1.06, pctdistance=0.76,
            wedgeprops={"width": 0.40, "edgecolor": "white", "linewidth": 2}
        )

        for wedge, (_, row) in zip(wedges, chart_data.iterrows()):
            logo_url = row.get("Logo", None)
            if logo_url is None or pd.isna(logo_url):
                continue
            try:
                img = self.chart_logo_cache.get(logo_url)
                if img is None:
                    continue
            except Exception:
                continue
            imagebox = OffsetImage(img, zoom=0.12)
            angle = (wedge.theta1 + wedge.theta2) / 2
            x = 1.1 * np.cos(np.deg2rad(angle))
            y = 1.1 * np.sin(np.deg2rad(angle))
            ab = AnnotationBbox(imagebox, (x, y), frameon=False)
            ax.add_artist(ab)

        for t in texts:
            t.set_visible(False)
            t.set_fontsize(13)
            t.set_fontweight("bold")

        for i, at in enumerate(autotexts):
            slice_color = slice_colors[i % len(slice_colors)]
            at.set_fontproperties(prop)
            at.set_fontsize(10)
            at.set_color(get_best_font_color(slice_color))
            at.set_fontweight("bold")

        # CENTER TEXT shows Total Seats and Total Valid Votes
        ax.text(0, 0.15, f"{total_seats}", ha="center", va="center", fontsize=30, fontweight="bold", color=PRIMARY_TEXT)
        ax.text(0, 0.0, "Total Seats", ha="center", va="center", fontsize=12, fontweight="bold", color=SUBTLE_TEXT)
        ax.text(0, -0.15, f"{int(self.total_valid_votes):,}", ha="center", va="center", fontsize=16, fontweight="bold", color=ACCENT_BLUE)
        ax.text(0, -0.25, "Total Eligible Votes", ha="center", va="center", fontsize=10, fontweight="bold", color=SUBTLE_TEXT)

        fig.suptitle("PR Seat Allocation Result", fontfamily=FONT_NAME, fontsize=20, fontweight="bold", y=0.99, color=PRIMARY_TEXT)
        ax.axis("equal")
        fig.subplots_adjust(top=0.92, bottom=0.04, left=0.04, right=0.96)

        fig.text(
            0.01,
            0.01,
            f"Source: {self.data_source}",
            ha="left",
            va="bottom",
            fontsize=10,
            color="#2c3e50",
            fontfamily=FONT_NAME,
            fontweight="bold",
        )

        # Pan/Zoom
        def on_scroll(event):
            if event.inaxes != ax:
                return
            base_scale = 1.1
            scale_factor = 1 / base_scale if event.button == 'up' else base_scale
            cur_xlim, cur_ylim = ax.get_xlim(), ax.get_ylim()
            new_w = (cur_xlim[1] - cur_xlim[0]) * scale_factor
            new_h = (cur_ylim[1] - cur_ylim[0]) * scale_factor
            rel_x = (cur_xlim[1] - event.xdata) / (cur_xlim[1] - cur_xlim[0])
            rel_y = (cur_ylim[1] - event.ydata) / (cur_ylim[1] - cur_ylim[0])
            ax.set_xlim([event.xdata - new_w * (1 - rel_x), event.xdata + new_w * rel_x])
            ax.set_ylim([event.ydata - new_h * (1 - rel_y), event.ydata + new_h * rel_y])
            fig.canvas.draw_idle()

        def on_press(event):
            if event.inaxes != ax:
                return
            self.pan_start = (event.xdata, event.ydata)

        def on_motion(event):
            if self.pan_start is None or event.inaxes != ax:
                return
            dx, dy = event.xdata - self.pan_start[0], event.ydata - self.pan_start[1]
            ax.set_xlim(ax.get_xlim() - dx)
            ax.set_ylim(ax.get_ylim() - dy)
            fig.canvas.draw_idle()

        fig.canvas.mpl_connect('scroll_event', on_scroll)
        fig.canvas.mpl_connect('button_press_event', on_press)
        fig.canvas.mpl_connect('motion_notify_event', on_motion)
        fig.canvas.mpl_connect('button_release_event', lambda e: setattr(self, 'pan_start', None))

        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    def process(self):
        if not self.filepath:
            return
        try:
            raw_df = pd.read_excel(self.filepath, header=None)
            df_clean = raw_df.dropna(how="all").reset_index(drop=True)

            manual_total = self.clean_number(df_clean.iloc[-1, 0])
            self.total_input_votes = manual_total

            party_data = []
            for i in range(1, len(df_clean) - 1, 2):
                p_name = str(df_clean.iloc[i, 0]).strip()
                p_votes = self.clean_number(df_clean.iloc[i + 1, 0])
                if p_name and p_votes > 0:
                    party_data.append({"Party": p_name, "Votes": p_votes})

            res, total_input, thresh, count, total_p, total_valid = calculate_seats(
                pd.DataFrame(party_data),
                int(self.seats_entry.get()),
                3.0,
                manual_total
            )
            self.result_df = res
            self.total_valid_votes = total_valid

            # Update summary labels (order fixed)
            self.lbl_total_input_votes.config(text=f"Total Votes: {int(self.total_input_votes):,}")
            self.lbl_threshold.config(text=f"3% Threshold: {int(thresh):,}")
            self.lbl_total_votes.config(text=f"Total Valid Votes: {int(self.total_valid_votes):,}")

            # Refresh qualified table
            for i in self.tree_qualified.get_children():
                self.tree_qualified.delete(i)
            if res is not None:
                for _, r in res.iterrows():
                    self.tree_qualified.insert("", tk.END, values=(r["Party"], f"{int(r['Votes']):,}", f" {int(r['Seats'])} "))

            self.update_chart()
            _beep()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def reset_app(self):
        self.filepath = None
        self.result_df = None
        self.total_input_votes = 0
        self.total_valid_votes = 0

        for i in self.tree_qualified.get_children():
            self.tree_qualified.delete(i)
        for w in self.chart_frame.winfo_children():
            w.destroy()
        self.clear_unqualified_grid()

        if self.current_figure:
            plt.close(self.current_figure)
            self.current_figure = None

        # Reset the three labels only
        self.lbl_total_input_votes.config(text="Total Votes: -")
        self.lbl_threshold.config(text="3% Threshold: -")
        self.lbl_total_votes.config(text="Total Valid Votes: -")
        _beep()


if __name__ == "__main__":
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(u"pumori.engineering.seatcalc.pro")  # type: ignore[attr-defined]
    except Exception:
        pass
    root = tk.Tk()
    app = ElectionApp(root)
    root.mainloop()
