from openpyxl import load_workbook
import pandas as pd
from scraper import fetch_pr_votes


def _headers_map(ws):
    """Return {header_text: 1-based column_index} from header row 1."""
    return {str(c.value).strip(): c.col_idx for c in ws[1] if c.value is not None}


def _find_sheet_with_headers(wb, required_headers):
    """Return (sheet_name, headers_map) for the first sheet having all required headers."""
    for name in wb.sheetnames:
        ws = wb[name]
        hdrs = _headers_map(ws)
        if all(h in hdrs for h in required_headers):
            return name, hdrs
    return None, None


def update_excel_from_web(
    file_path: str,
    sheet_name: str | None = None,
    party_header: str = "Party",
    votes_header: str = "Votes",
    logo_url_header: str = "Logo_URL",
):
    """
    Update ONLY Party, Votes, Logo_URL values in an existing workbook.
    Does not modify formulas, formatting, widths, hidden columns, etc.

    Returns:
        dict -> {'sheet', 'updated', 'not_found'}
    """

    wb = load_workbook(file_path)

    # Resolve sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = _headers_map(ws)
        for h in (party_header, votes_header, logo_url_header):
            if h not in headers:
                raise ValueError(f"Sheet '{sheet_name}' missing required column '{h}'.")
    else:
        name, headers = _find_sheet_with_headers(
            wb, (party_header, votes_header, logo_url_header)
        )
        if not name:
            raise ValueError(
                f"No sheet containing required headers: {party_header}, {votes_header}, {logo_url_header}"
            )
        ws = wb[name]

    party_col = headers[party_header]
    votes_col = headers[votes_header]
    logo_col = headers[logo_url_header]

    # Fetch latest PR data
    df, _ = fetch_pr_votes()

    if not {"Party", "Votes", "Logo"}.issubset(df.columns):
        raise ValueError("fetch_pr_votes() must return columns: Party, Votes, Logo")

    def norm(x):
        return "" if x is None else str(x).strip()

    lookup = {
        norm(p): (int(v), norm(l))
        for p, v, l in zip(df["Party"], df["Votes"], df["Logo"])
    }

    updated = 0
    not_found = []

    # Sort data by votes descending (same as website)
    df = df.sort_values("Votes", ascending=False).reset_index(drop=True)

    start_row = 2
    
    if len(df) > ws.max_row - 1:
        raise ValueError("Excel sheet does not have enough rows for all parties.")

    for i, row in df.iterrows():

        excel_row = start_row + i

        ws.cell(row=excel_row, column=party_col).value = str(row["Party"]).strip()
        ws.cell(row=excel_row, column=votes_col).value = int(row["Votes"])
        ws.cell(row=excel_row, column=logo_col).value = str(row["Logo"]).strip()

    updated = len(df)

    wb.save(file_path)

    return {
        "sheet": ws.title,
        "updated": updated,
        "not_found": not_found,
    }


def read_pr_from_excel(file_path):

    df = pd.read_excel(file_path)

    if "Party" not in df.columns or "Votes" not in df.columns:
        raise ValueError("Excel must contain columns: Party and Votes")

    # Clean Party names
    df["Party"] = df["Party"].astype(str).str.strip()
    df = df[df["Party"] != ""]

    # Clean Votes
    df["Votes"] = (
        df["Votes"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("\xa0", "", regex=False)
        .str.strip()
    )

    df["Votes"] = pd.to_numeric(df["Votes"], errors="coerce").fillna(0).astype(int)

    # Handle Logo column
    if "Logo_URL" in df.columns:
        df = df.rename(columns={"Logo_URL": "Logo"})
    elif "Logo" not in df.columns:
        df["Logo"] = None

    df = df[["Party", "Votes", "Logo"]].copy()

    df["Logo"] = df["Logo"].where(pd.notna(df["Logo"]), None)

    total_votes = int(df["Votes"].sum())

    return df, total_votes
